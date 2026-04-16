"""Convert raw eFIP spreadsheets into the JSONL files used by the pipeline."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .io_utils import write_jsonl


@dataclass(frozen=True)
class SentenceMatch:
    """A matched sentence plus the spans used for marker insertion."""

    sentence_id: str
    text: str
    first_entity: str
    first_span: tuple[int, int]
    second_entity: str
    second_span: tuple[int, int]


@dataclass(frozen=True)
class ConversionSummary:
    """Summary counts for an eFIP conversion run."""

    input_rows: int
    output_rows: int
    skipped_rows: int
    unsupported_overlap_rows: int
    unmatched_rows: int


@dataclass(frozen=True)
class EfipConversionResult:
    """Converted rows and stats for both raw eFIP sources."""

    corpus_rows: list[dict]
    full_rows: list[dict]
    multi_sentence_examples: list[dict]
    corpus_summary: ConversionSummary
    full_summary: ConversionSummary


def normalize_text(value: Any) -> str:
    """Normalize a cell value into a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def workbook_rows(path: Path) -> list[dict[str, Any]]:
    """Load the first worksheet of an XLSX file into a list of dictionaries."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    results: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {}
        for index, column_name in enumerate(header):
            if column_name is not None:
                item[str(column_name)] = row[index] if index < len(row) else None
        results.append(item)
    return results


def sentence_candidates_for_entity(text: str, entity: str) -> list[tuple[int, int]]:
    """Find case-insensitive candidate spans for an entity string in text."""
    cleaned = normalize_text(entity)
    if not cleaned:
        return []

    tokens = re.split(r"\s+", cleaned)
    pattern = r"\s+".join(re.escape(token) for token in tokens if token)
    return [(match.start(), match.end()) for match in re.finditer(pattern, text, flags=re.IGNORECASE)]


def choose_entity_pair(text: str, left_entity: str, right_entity: str) -> SentenceMatch | None:
    """Choose the first non-overlapping entity pair in a sentence."""
    left_matches = sentence_candidates_for_entity(text, left_entity)
    right_matches = sentence_candidates_for_entity(text, right_entity)
    if not left_matches or not right_matches:
        return None

    for left_span in left_matches:
        for right_span in right_matches:
            if left_span == right_span:
                continue
            if left_span[1] <= right_span[0] or right_span[1] <= left_span[0]:
                if left_span[0] <= right_span[0]:
                    return SentenceMatch(
                        sentence_id="",
                        text=text,
                        first_entity=normalize_text(left_entity),
                        first_span=left_span,
                        second_entity=normalize_text(right_entity),
                        second_span=right_span,
                    )
                return SentenceMatch(
                    sentence_id="",
                    text=text,
                    first_entity=normalize_text(right_entity),
                    first_span=right_span,
                    second_entity=normalize_text(left_entity),
                    second_span=left_span,
                )
    return None


def insert_markers(text: str, first_span: tuple[int, int], second_span: tuple[int, int]) -> tuple[str, tuple[int, int], tuple[int, int]]:
    """Insert [E1] and [E2] markers into a sentence."""
    first_start, first_end = first_span
    second_start, second_end = second_span

    prefix = text[:first_start]
    first_text = text[first_start:first_end]
    middle = text[first_end:second_start]
    second_text = text[second_start:second_end]
    suffix = text[second_end:]

    marked_text = f"{prefix}[E1]{first_text}[/E1]{middle}[E2]{second_text}[/E2]{suffix}"

    first_marked_span = (len(prefix) + 4, len(prefix) + 4 + len(first_text))
    second_marked_start = len(prefix) + 4 + len(first_text) + 5 + len(middle) + 4
    second_marked_span = (second_marked_start, second_marked_start + len(second_text))
    return marked_text, first_marked_span, second_marked_span


def build_relation(
    first_entity: str,
    first_span: tuple[int, int],
    first_marked_span: tuple[int, int],
    second_entity: str,
    second_span: tuple[int, int],
    second_marked_span: tuple[int, int],
    relation_type: str,
) -> list[dict]:
    """Build the normalized relation list for one output row."""
    return [
        {
            "PPI_relation_type": normalize_text(relation_type),
            "relation_id": 0,
            "entity_1": first_entity,
            "entity_1_idx": [[first_span[0], first_span[1]]],
            "entity_1_idx_in_text_with_entity_marker": [first_marked_span[0], first_marked_span[1]],
            "entity_1_type": "protein",
            "entity_1_type_id": 0,
            "entity_2": second_entity,
            "entity_2_idx": [[second_span[0], second_span[1]]],
            "entity_2_idx_in_text_with_entity_marker": [second_marked_span[0], second_marked_span[1]],
            "entity_2_type": "protein",
            "entity_2_type_id": 0,
        }
    ]


class EfipCorpusSentenceStore:
    """Lazy loader for the subsection workbooks used by the eFIP corpus annotations."""

    def __init__(self, subsections_dir: Path) -> None:
        self.subsections_dir = subsections_dir
        self._cache: dict[str, list[dict[str, Any]]] = {}

    def rows_for_pmc(self, pmc_id: str) -> list[dict[str, Any]]:
        if pmc_id not in self._cache:
            rows: list[dict[str, Any]] = []
            for path in sorted(self.subsections_dir.glob(f"{pmc_id}*.xlsx")):
                rows.extend(workbook_rows(path))
            self._cache[pmc_id] = rows
        return self._cache[pmc_id]


def parse_sent_id_tokens(value: Any) -> list[str]:
    """Split a SentIDs cell into ordered sentence-id tokens."""
    if value is None:
        return []
    if isinstance(value, (int, float)) and int(value) == value:
        return [str(int(value))]
    return [token.strip() for token in str(value).split(",") if token.strip()]


def section_matches_subsec(section_value: Any, subsec_id: str) -> bool:
    """Return True when a subsection row belongs to the desired subsection id."""
    section = normalize_text(section_value)
    return section.endswith(subsec_id)


def candidate_sentences_for_annotation(
    sentence_store: EfipCorpusSentenceStore,
    pmc_id: str,
    subsec_id: str,
    sent_ids: Any,
) -> list[tuple[str, str]]:
    """Resolve raw corpus annotation sentence ids to actual sentence text."""
    rows = sentence_store.rows_for_pmc(pmc_id)
    tokens = parse_sent_id_tokens(sent_ids)
    candidates: list[tuple[str, str]] = []

    for token in tokens:
        matched = []
        for row in rows:
            row_sent_id = normalize_text(row.get("SentID"))
            if row_sent_id != token:
                continue
            if token == "-1" or section_matches_subsec(row.get("Section"), subsec_id):
                matched.append((token, normalize_text(row.get("Sentence"))))
        if not matched:
            for row in rows:
                row_sent_id = normalize_text(row.get("SentID"))
                if row_sent_id == token:
                    matched.append((token, normalize_text(row.get("Sentence"))))
        candidates.extend(matched)

    return candidates


def build_corpus_row(raw_row: dict[str, Any], match: SentenceMatch) -> dict:
    """Build one converted corpus JSON row."""
    marked_text, first_marked_span, second_marked_span = insert_markers(match.text, match.first_span, match.second_span)
    return {
        "id": f"eFIP_Corpus_PMC{normalize_text(raw_row['PMC ID'])}_S{match.sentence_id}",
        "text": match.text,
        "text_with_entity_marker": marked_text,
        "relation": build_relation(
            match.first_entity,
            match.first_span,
            first_marked_span,
            match.second_entity,
            match.second_span,
            second_marked_span,
            normalize_text(raw_row["PPI"]),
        ),
        "PMC ID": int(float(raw_row["PMC ID"])) if raw_row.get("PMC ID") not in (None, "") else None,
        "Subsec ID": int(float(raw_row["Subsec ID"])) if raw_row.get("Subsec ID") not in (None, "") else None,
        "Subsec Type": normalize_text(raw_row["Subsec Type"]).title() if raw_row.get("Subsec Type") not in (None, "") else None,
        "Kinase": raw_row.get("Kinase"),
        "Substrate": raw_row.get("Substrate"),
        "Site": raw_row.get("Site"),
        "Impact": raw_row.get("Impact"),
        "PPI": raw_row.get("PPI"),
        "Interactant": raw_row.get("Interactant"),
        "SentIDs": raw_row.get("SentIDs"),
    }


def build_full_row(raw_row: dict[str, Any], match: SentenceMatch) -> dict:
    """Build one converted sentence-level eFIP JSON row."""
    marked_text, first_marked_span, second_marked_span = insert_markers(match.text, match.first_span, match.second_span)
    return {
        "id": f"eFIP_PMID{int(raw_row['PMID'])}_S{int(raw_row['Sentence ID'])}",
        "text": match.text,
        "text_with_entity_marker": marked_text,
        "relation": build_relation(
            match.first_entity,
            match.first_span,
            first_marked_span,
            match.second_entity,
            match.second_span,
            second_marked_span,
            normalize_text(raw_row["PPI Type (verb)"]),
        ),
        "Relevancy": raw_row.get("Relevancy"),
        "PMID": raw_row.get("PMID"),
        "Phospho-protein": raw_row.get("Phospho-protein"),
        "Phospho-site": raw_row.get("Phospho-site"),
        "Kinase": raw_row.get("Kinase"),
        "Interactant": raw_row.get("Interactant"),
        "PPI Type (verb)": raw_row.get("PPI Type (verb)"),
        "Effect (modifier)": raw_row.get("Effect (modifier)"),
        "Sentence ID": raw_row.get("Sentence ID"),
        "Sentence": raw_row.get("Sentence"),
    }


def convert_efip_corpus(annotations_path: Path, subsections_dir: Path, output_path: Path) -> tuple[list[dict], list[dict], ConversionSummary]:
    """Convert the raw eFIP corpus annotations into JSONL."""
    raw_rows = workbook_rows(annotations_path)
    sentence_store = EfipCorpusSentenceStore(subsections_dir)
    converted_rows: list[dict] = []
    multi_sentence_examples: list[dict] = []
    skipped_rows = 0
    unsupported_overlap_rows = 0
    unmatched_rows = 0

    for raw_row in raw_rows:
        pmc_id = normalize_text(raw_row.get("PMC ID"))
        subsec_id = normalize_text(raw_row.get("Subsec ID"))
        substrate = normalize_text(raw_row.get("Substrate"))
        interactant = normalize_text(raw_row.get("Interactant"))
        if not pmc_id or not substrate or not interactant:
            skipped_rows += 1
            continue

        sentence_hits = 0
        saw_overlap_only = False
        for sentence_id, sentence_text in candidate_sentences_for_annotation(
            sentence_store,
            pmc_id,
            subsec_id,
            raw_row.get("SentIDs"),
        ):
            match = choose_entity_pair(sentence_text, substrate, interactant)
            if match is None:
                if sentence_candidates_for_entity(sentence_text, substrate) and sentence_candidates_for_entity(sentence_text, interactant):
                    saw_overlap_only = True
                continue

            sentence_hits += 1
            converted_rows.append(
                build_corpus_row(
                    raw_row,
                    SentenceMatch(
                        sentence_id=sentence_id,
                        text=sentence_text,
                        first_entity=match.first_entity,
                        first_span=match.first_span,
                        second_entity=match.second_entity,
                        second_span=match.second_span,
                    ),
                )
            )

        if sentence_hits == 0:
            skipped_rows += 1
            if saw_overlap_only:
                unsupported_overlap_rows += 1
            else:
                unmatched_rows += 1
        elif len(parse_sent_id_tokens(raw_row.get("SentIDs"))) > 1 and len(multi_sentence_examples) < 20:
            for row in converted_rows[-sentence_hits:]:
                multi_sentence_examples.append(row)

    write_jsonl(output_path, converted_rows)
    return (
        converted_rows,
        multi_sentence_examples,
        ConversionSummary(
            input_rows=len(raw_rows),
            output_rows=len(converted_rows),
            skipped_rows=skipped_rows,
            unsupported_overlap_rows=unsupported_overlap_rows,
            unmatched_rows=unmatched_rows,
        ),
    )


def convert_efip_full(input_path: Path, output_path: Path) -> tuple[list[dict], ConversionSummary]:
    """Convert the sentence-level eFIP spreadsheet into JSONL."""
    raw_rows = workbook_rows(input_path)
    converted_rows: list[dict] = []
    skipped_rows = 0
    unsupported_overlap_rows = 0
    unmatched_rows = 0

    for raw_row in raw_rows:
        substrate = normalize_text(raw_row.get("Phospho-protein"))
        interactant = normalize_text(raw_row.get("Interactant"))
        sentence = normalize_text(raw_row.get("Sentence"))
        if not substrate or not interactant or not sentence:
            skipped_rows += 1
            continue

        match = choose_entity_pair(sentence, substrate, interactant)
        if match is None:
            skipped_rows += 1
            if sentence_candidates_for_entity(sentence, substrate) and sentence_candidates_for_entity(sentence, interactant):
                unsupported_overlap_rows += 1
            else:
                unmatched_rows += 1
            continue

        converted_rows.append(
            build_full_row(
                raw_row,
                SentenceMatch(
                    sentence_id=str(raw_row["Sentence ID"]),
                    text=sentence,
                    first_entity=match.first_entity,
                    first_span=match.first_span,
                    second_entity=match.second_entity,
                    second_span=match.second_span,
                ),
            )
        )

    write_jsonl(output_path, converted_rows)
    return (
        converted_rows,
        ConversionSummary(
            input_rows=len(raw_rows),
            output_rows=len(converted_rows),
            skipped_rows=skipped_rows,
            unsupported_overlap_rows=unsupported_overlap_rows,
            unmatched_rows=unmatched_rows,
        ),
    )


def convert_efip_sources(
    corpus_annotations_path: Path,
    corpus_subsections_dir: Path,
    full_input_path: Path,
    corpus_output_path: Path,
    full_output_path: Path,
    multi_sentence_sample_path: Path | None = None,
) -> EfipConversionResult:
    """Convert both raw eFIP sources into the normalized JSON outputs."""
    corpus_rows, multi_sentence_examples, corpus_summary = convert_efip_corpus(
        corpus_annotations_path,
        corpus_subsections_dir,
        corpus_output_path,
    )
    full_rows, full_summary = convert_efip_full(full_input_path, full_output_path)

    if multi_sentence_sample_path is not None:
        multi_sentence_sample_path.parent.mkdir(parents=True, exist_ok=True)
        multi_sentence_sample_path.write_text(json.dumps(multi_sentence_examples, ensure_ascii=False, indent=2), encoding="utf-8")

    return EfipConversionResult(
        corpus_rows=corpus_rows,
        full_rows=full_rows,
        multi_sentence_examples=multi_sentence_examples,
        corpus_summary=corpus_summary,
        full_summary=full_summary,
    )
